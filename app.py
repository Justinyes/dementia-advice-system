from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - app still explains the dependency issue.
    Workbook = None
    load_workbook = None


ROOT = Path(__file__).parent.resolve()
DATA_DIR = ROOT / "data"
EVIDENCE_PATH = DATA_DIR / "evidence.xlsx"


KEY_FIELDS = [
    "age",
    "chronic_conditions",
    "mobility_risk",
    "sleep",
    "social_activity",
    "sensory_status",
]


FIELD_LABELS = {
    "age": "年龄",
    "relationship": "对象关系",
    "chronic_conditions": "慢病情况",
    "mobility_risk": "运动限制/跌倒风险",
    "sleep": "睡眠",
    "social_activity": "社交参与",
    "sensory_status": "听力/视力",
    "preferences": "偏好",
    "red_flags": "红旗信号",
}


SAMPLE_EVIDENCE = [
    {
        "id": "EV-001",
        "intervention": "关节友好的低强度有氧 + 力量/平衡训练",
        "category": "运动",
        "applicable_tags": "older_adult,low_activity,hypertension,knee_pain,cardiometabolic_risk",
        "outcome_tags": "cognitive_decline,dementia_risk,function",
        "contraindication_tags": "unstable_chest_pain,acute_injury,severe_fall_risk",
        "evidence_level": "A",
        "recommendation_strength": "strong",
        "summary": "规律身体活动与较低的认知下降风险相关；对有膝痛者应优先选择低冲击形式，并加入平衡与力量训练以降低跌倒风险。",
        "source_id": "SRC-001",
        "source_citation": "WHO risk reduction of cognitive decline and dementia guideline; multidomain lifestyle intervention reviews.",
        "action_template": "从每周 3 次、每次 15-20 分钟开始，可选快走、固定车、八段锦、坐站训练和简化平衡训练。",
        "safety_note": "若出现胸痛、明显气短、头晕、近期跌倒或关节疼痛加重，应先咨询医生或康复专业人员。",
    },
    {
        "id": "EV-002",
        "intervention": "认知 + 社交活动，麻将/牌类可作为形式之一",
        "category": "认知与社交",
        "applicable_tags": "social_low,mahjong_preference,willing_go_out,loneliness,older_adult",
        "outcome_tags": "cognitive_stimulation,mood,social_participation",
        "contraindication_tags": "gambling_conflict,severe_behavior_problem",
        "evidence_level": "B",
        "recommendation_strength": "moderate",
        "summary": "持续社交参与和认知刺激活动与更好的认知和情绪状态相关；熟悉且愿意参与的活动更容易坚持。",
        "source_id": "SRC-002",
        "source_citation": "Cognitive stimulation and social participation evidence summaries in dementia prevention literature.",
        "action_template": "保留每周 2-4 次熟人麻将或牌类活动，控制时长，避免赌博化，并搭配散步、聊天或学习新规则。",
        "safety_note": "如活动引发争执、熬夜、久坐过长或金钱风险，应调整为非赌博、短时段、有人陪伴的形式。",
    },
    {
        "id": "EV-003",
        "intervention": "血压/血糖/血脂等慢病管理与随访",
        "category": "慢病管理",
        "applicable_tags": "hypertension,diabetes,dyslipidemia,cardiometabolic_risk,older_adult",
        "outcome_tags": "vascular_risk,cognitive_decline,dementia_risk",
        "contraindication_tags": "",
        "evidence_level": "A",
        "recommendation_strength": "strong",
        "summary": "心血管和代谢风险管理是降低认知下降风险的重要组成部分，尤其适用于已有高血压、糖尿病或血脂异常的人群。",
        "source_id": "SRC-003",
        "source_citation": "Lancet Commission dementia prevention reports; WHO cognitive decline risk reduction guideline.",
        "action_template": "记录家庭血压，按医嘱复诊，和医生确认血压、血糖、血脂目标及用药依从性。",
        "safety_note": "不要自行增减处方药；头晕、晕厥、胸痛或血压异常波动时应及时就医。",
    },
    {
        "id": "EV-004",
        "intervention": "睡眠节律与失眠行为干预",
        "category": "睡眠",
        "applicable_tags": "sleep_poor,sleep_average,daytime_sleepiness,older_adult",
        "outcome_tags": "sleep_quality,mood,cognitive_function",
        "contraindication_tags": "suspected_sleep_apnea",
        "evidence_level": "B",
        "recommendation_strength": "moderate",
        "summary": "睡眠质量与认知、情绪和白天功能相关；稳定作息和睡眠卫生是低风险的基础干预。",
        "source_id": "SRC-004",
        "source_citation": "Sleep health and cognitive aging reviews; behavioral sleep intervention guidance.",
        "action_template": "固定起床时间，午睡控制在 30 分钟以内，下午减少浓茶咖啡，睡前 1 小时减少屏幕刺激。",
        "safety_note": "若有严重打鼾、憋醒、白天嗜睡或长期失眠，应到睡眠门诊评估。",
    },
    {
        "id": "EV-005",
        "intervention": "听力/视力筛查与矫正",
        "category": "感官干预",
        "applicable_tags": "hearing_problem,vision_problem,unknown_sensory,older_adult",
        "outcome_tags": "communication,social_participation,cognitive_load",
        "contraindication_tags": "",
        "evidence_level": "B",
        "recommendation_strength": "moderate",
        "summary": "听力和视力问题会影响交流、社交参与和日常功能；筛查和矫正是痴呆风险管理中的重要可改变因素。",
        "source_id": "SRC-005",
        "source_citation": "Lancet Commission dementia prevention reports; sensory impairment and cognitive aging reviews.",
        "action_template": "安排听力和视力检查，必要时验配助听器、眼镜或处理白内障等可矫正问题。",
        "safety_note": "突然听力或视力下降属于需要及时就医的情况。",
    },
    {
        "id": "EV-006",
        "intervention": "饮食结构优化：少超加工，增加蔬果、鱼类、豆类和坚果",
        "category": "饮食",
        "applicable_tags": "hypertension,diabetes,dyslipidemia,cardiometabolic_risk,older_adult",
        "outcome_tags": "vascular_risk,dementia_risk,general_health",
        "contraindication_tags": "renal_diet_restriction,swallowing_problem",
        "evidence_level": "B",
        "recommendation_strength": "moderate",
        "summary": "接近地中海式或均衡膳食模式与较好的心血管和认知健康相关；减少超加工食品有助于整体风险管理。",
        "source_id": "SRC-006",
        "source_citation": "Mediterranean/MIND diet and cognitive aging systematic reviews.",
        "action_template": "每餐有蔬菜，优先鱼、蛋、豆制品和全谷物，少油炸、腌制和高糖零食。",
        "safety_note": "若有肾病、吞咽困难或糖尿病饮食医嘱，应按医生/营养师建议调整。",
    },
]


@dataclass
class ProfileItem:
    status: str
    value: Any
    evidence: str
    confidence: float


@dataclass
class SessionState:
    messages: list[dict[str, str]] = field(default_factory=list)
    profile: dict[str, ProfileItem] = field(default_factory=dict)
    candidates: dict[str, ProfileItem] = field(default_factory=dict)


SESSIONS: dict[str, SessionState] = {"default": SessionState()}


def ensure_evidence_workbook() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    if EVIDENCE_PATH.exists():
        return
    if Workbook is None:
        raise RuntimeError("openpyxl 不可用，无法创建 Excel 证据库。")
    wb = Workbook()
    ws = wb.active
    ws.title = "evidence"
    headers = list(SAMPLE_EVIDENCE[0].keys())
    ws.append(headers)
    for item in SAMPLE_EVIDENCE:
        ws.append([item[h] for h in headers])
    source = wb.create_sheet("sources")
    source.append(["source_id", "source_citation"])
    seen = set()
    for item in SAMPLE_EVIDENCE:
        if item["source_id"] not in seen:
            source.append([item["source_id"], item["source_citation"]])
            seen.add(item["source_id"])
    wb.save(EVIDENCE_PATH)


def load_evidence() -> list[dict[str, str]]:
    ensure_evidence_workbook()
    if load_workbook is None:
        raise RuntimeError("openpyxl 不可用，无法读取 Excel 证据库。")
    wb = load_workbook(EVIDENCE_PATH, read_only=True, data_only=True)
    ws = wb["evidence"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v) for v in rows[0]]
    evidence = []
    for row in rows[1:]:
        evidence.append({headers[i]: "" if row[i] is None else str(row[i]) for i in range(len(headers))})
    return evidence


def split_tags(value: str) -> set[str]:
    return {tag.strip() for tag in value.split(",") if tag.strip()}


def put_update(updates: dict[str, ProfileItem], field_name: str, value: Any, evidence: str, confidence: float = 0.9) -> None:
    status = "known" if confidence >= 0.7 else "uncertain"
    updates[field_name] = ProfileItem(status=status, value=value, evidence=evidence, confidence=confidence)


def extract_profile_updates(message: str) -> tuple[dict[str, ProfileItem], set[str]]:
    text = message.strip()
    updates: dict[str, ProfileItem] = {}
    tags: set[str] = set()

    age_match = re.search(r"(\d{2,3})\s*岁", text)
    if age_match:
        age = int(age_match.group(1))
        put_update(updates, "age", age, age_match.group(0))
        if age >= 60:
            tags.add("older_adult")

    if any(word in text for word in ["我爸", "父亲", "爸爸"]):
        put_update(updates, "relationship", "父亲", "提到父亲/爸爸")
    elif any(word in text for word in ["我妈", "母亲", "妈妈"]):
        put_update(updates, "relationship", "母亲", "提到母亲/妈妈")
    elif "我" in text:
        put_update(updates, "relationship", "本人", "用户以第一人称描述")

    chronic = []
    if "高血压" in text or "血压高" in text:
        chronic.append("高血压")
        tags.update(["hypertension", "cardiometabolic_risk"])
    if "糖尿病" in text or "血糖" in text:
        chronic.append("糖尿病")
        tags.update(["diabetes", "cardiometabolic_risk"])
    if "血脂" in text or "高脂" in text:
        chronic.append("血脂异常")
        tags.update(["dyslipidemia", "cardiometabolic_risk"])
    if chronic:
        put_update(updates, "chronic_conditions", chronic, "、".join(chronic))

    mobility = []
    if any(word in text for word in ["膝盖疼", "膝痛", "腿疼", "关节疼"]):
        mobility.append("膝/关节疼痛")
        tags.add("knee_pain")
    if any(word in text for word in ["跌倒", "摔倒", "站不稳"]):
        mobility.append("跌倒或平衡风险")
        tags.add("severe_fall_risk")
    if any(word in text for word in ["不爱运动", "不太爱运动", "很少运动", "久坐", "不怎么动"]):
        mobility.append("运动不足")
        tags.add("low_activity")
    if mobility:
        put_update(updates, "mobility_risk", mobility, "、".join(mobility))

    if any(word in text for word in ["失眠", "睡不好", "睡眠差", "睡得差"]):
        put_update(updates, "sleep", "睡眠差", "提到睡眠差/失眠")
        tags.add("sleep_poor")
    elif any(word in text for word in ["睡眠一般", "睡得一般"]):
        put_update(updates, "sleep", "睡眠一般", "提到睡眠一般")
        tags.add("sleep_average")

    if any(word in text for word in ["麻将", "打牌", "牌友"]):
        put_update(updates, "preferences", "喜欢牌类/麻将社交", "提到麻将/打牌")
        tags.add("mahjong_preference")
    if any(word in text for word in ["不出门", "社交少", "孤独", "很少见人"]):
        put_update(updates, "social_activity", "社交参与少", "提到社交少/孤独")
        tags.update(["social_low", "loneliness"])
    elif any(word in text for word in ["愿意出门", "邻居", "朋友", "社区"]):
        put_update(updates, "social_activity", "有外出或社交意愿", "提到出门/邻居/朋友")
        tags.add("willing_go_out")

    sensory = []
    sensory_unknown = any(word in text for word in ["没检查", "未检查", "没查过", "还没检查"])
    mentions_hearing_or_vision = any(word in text for word in ["听力", "视力", "耳", "眼"])
    if sensory_unknown and mentions_hearing_or_vision:
        sensory.append("听力/视力未检查")
        tags.add("unknown_sensory")
    elif any(word in text for word in ["听不清", "耳背"]):
        sensory.append("可能有听力问题")
        tags.add("hearing_problem")
    if not sensory_unknown and any(word in text for word in ["看不清", "眼睛模糊", "视物模糊"]):
        sensory.append("可能有视力问题")
        tags.add("vision_problem")
    if sensory:
        put_update(updates, "sensory_status", sensory, "、".join(sensory))

    red_flags = []
    if any(word in text for word in ["突然", "快速下降", "短时间", "明显变差"]):
        red_flags.append("短期内明显/快速下降")
    if any(word in text for word in ["自杀", "不想活", "活不下去", "伤害自己"]):
        red_flags.append("自伤或严重抑郁风险")
    if any(word in text for word in ["生活不能自理", "走失", "认不出家人", "幻觉"]):
        red_flags.append("生活功能或精神行为红旗")
    if red_flags:
        put_update(updates, "red_flags", red_flags, "、".join(red_flags), confidence=0.95)

    return updates, tags


def merge_updates(state: SessionState, updates: dict[str, ProfileItem]) -> None:
    for key, update in updates.items():
        if update.status == "known":
            state.profile[key] = update
        else:
            state.candidates[key] = update


def profile_tags(state: SessionState, new_tags: set[str]) -> set[str]:
    tags = set(new_tags)
    profile = state.profile
    if profile.get("age") and int(profile["age"].value) >= 60:
        tags.add("older_adult")
    for condition in profile.get("chronic_conditions", ProfileItem("unknown", [], "", 0)).value or []:
        if condition == "高血压":
            tags.update(["hypertension", "cardiometabolic_risk"])
        if condition == "糖尿病":
            tags.update(["diabetes", "cardiometabolic_risk"])
        if condition == "血脂异常":
            tags.update(["dyslipidemia", "cardiometabolic_risk"])
    for item in profile.get("mobility_risk", ProfileItem("unknown", [], "", 0)).value or []:
        if "膝" in item or "关节" in item:
            tags.add("knee_pain")
        if "跌倒" in item:
            tags.add("severe_fall_risk")
        if "运动不足" in item:
            tags.add("low_activity")
    sleep = profile.get("sleep")
    if sleep and sleep.value == "睡眠差":
        tags.add("sleep_poor")
    elif sleep and sleep.value == "睡眠一般":
        tags.add("sleep_average")
    social = profile.get("social_activity")
    if social and "少" in str(social.value):
        tags.add("social_low")
    elif social:
        tags.add("willing_go_out")
    pref = profile.get("preferences")
    if pref and "麻将" in str(pref.value):
        tags.add("mahjong_preference")
    if "sensory_status" not in profile:
        tags.add("unknown_sensory")
    return tags


def missing_fields(state: SessionState) -> list[str]:
    return [field for field in KEY_FIELDS if field not in state.profile]


def next_question(state: SessionState) -> str:
    missing = missing_fields(state)
    questions = {
        "age": "请问对象大概多大年龄？是本人、父亲还是母亲？",
        "chronic_conditions": "是否有高血压、糖尿病、血脂异常或其他慢病？",
        "mobility_risk": "平时运动情况如何？有没有膝痛、跌倒、胸闷气短或其他运动禁忌？",
        "sleep": "睡眠怎么样？是否失眠、打鼾憋醒或白天嗜睡？",
        "social_activity": "平时社交和外出多吗？有没有喜欢的活动，比如麻将、跳舞、散步、学习班？",
        "sensory_status": "听力和视力怎么样？有没有听不清、看不清或没做过检查？",
    }
    selected = missing[:2]
    return "我还需要补充两点，才能更稳妥地排序建议：\n" + "\n".join(f"- {questions[item]}" for item in selected)


def has_red_flags(state: SessionState) -> bool:
    return "red_flags" in state.profile


def recommendation_score(item: dict[str, str], tags: set[str]) -> tuple[int, list[str], bool]:
    evidence_points = {"A": 40, "B": 30, "C": 18}.get(item["evidence_level"].upper(), 10)
    strength_points = {"strong": 15, "moderate": 10, "weak": 5}.get(item["recommendation_strength"], 5)
    applicable = split_tags(item["applicable_tags"])
    contraindications = split_tags(item["contraindication_tags"])
    matched = applicable & tags
    blocked = bool(contraindications & tags)
    score = evidence_points + strength_points + len(matched) * 10
    reasons = [f"证据等级 {item['evidence_level']}，推荐强度 {item['recommendation_strength']}"]
    if matched:
        reasons.append("匹配画像标签：" + "、".join(sorted(matched)))
    if blocked:
        score -= 45
        reasons.append("存在需谨慎处理的风险标签：" + "、".join(sorted(contraindications & tags)))
    return score, reasons, blocked


def build_recommendations(state: SessionState, tags: set[str]) -> list[dict[str, Any]]:
    scored = []
    for item in load_evidence():
        score, reasons, blocked = recommendation_score(item, tags)
        if score < 35:
            continue
        scored.append(
            {
                **item,
                "score": score,
                "reasons": reasons,
                "blocked": blocked,
            }
        )
    scored.sort(key=lambda row: row["score"], reverse=True)
    return scored[:5]


def profile_as_json(state: SessionState) -> list[dict[str, Any]]:
    result = []
    for key, item in state.profile.items():
        result.append(
            {
                "field": key,
                "label": FIELD_LABELS.get(key, key),
                "status": item.status,
                "value": item.value,
                "evidence": item.evidence,
                "confidence": item.confidence,
            }
        )
    return result


def advice_text(recommendations: list[dict[str, Any]], state: SessionState) -> str:
    if has_red_flags(state):
        flags = "、".join(state.profile["red_flags"].value)
        lines = [
            f"你提到的情况包含红旗信号：{flags}。",
            "\n建议尽快到记忆门诊、神经内科、老年医学科或精神心理相关门诊做系统评估。下面的排序只作为等待评估期间的保守生活管理参考，不替代诊断或治疗。",
        ]
    else:
        lines = ["根据目前画像和证据库，优先级建议如下："]

    for idx, item in enumerate(recommendations, start=1):
        lines.append(
            f"\n{idx}. {item['intervention']}（评分 {item['score']}）\n"
            f"为什么适合：{'；'.join(item['reasons'])}。\n"
            f"怎么做：{item['action_template']}\n"
            f"证据摘要：{item['summary']}\n"
            f"来源：{item['source_citation']}\n"
            f"安全提示：{item['safety_note']}"
        )
    lines.append("\n免责声明：以上内容仅用于健康教育和生活方式建议，不构成诊断、处方或替代医生面诊。")
    return "\n".join(lines)


def handle_chat(payload: dict[str, Any]) -> dict[str, Any]:
    session_id = payload.get("session_id", "default")
    message = str(payload.get("message", "")).strip()
    state = SESSIONS.setdefault(session_id, SessionState())
    if not message:
        return {"error": "message 不能为空"}

    state.messages.append({"role": "user", "content": message})
    updates, tags = extract_profile_updates(message)
    merge_updates(state, updates)
    tags = profile_tags(state, tags)

    if has_red_flags(state):
        recommendations = build_recommendations(state, tags)
        reply = advice_text(recommendations, state)
        phase = "red_flag"
    elif len(missing_fields(state)) > 2:
        recommendations = []
        reply = next_question(state)
        phase = "question"
    else:
        recommendations = build_recommendations(state, tags)
        reply = advice_text(recommendations, state)
        phase = "recommend"

    state.messages.append({"role": "assistant", "content": reply})
    return {
        "reply": reply,
        "phase": phase,
        "profile": profile_as_json(state),
        "missing_fields": missing_fields(state),
        "tags": sorted(tags),
        "recommendations": recommendations,
    }


class Handler(SimpleHTTPRequestHandler):
    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/profile":
            self.write_json({"profile": profile_as_json(SESSIONS["default"])})
            return
        if parsed.path == "/":
            self.path = "/static/index.html"
        return super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/api/chat":
            self.send_error(404)
            return
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        try:
            payload = json.loads(body.decode("utf-8"))
            result = handle_chat(payload)
            self.write_json(result)
        except Exception as exc:
            self.write_json({"error": str(exc)}, status=500)

    def write_json(self, data: dict[str, Any], status: int = 200) -> None:
        encoded = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)


def main() -> None:
    os.chdir(ROOT)
    ensure_evidence_workbook()
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"痴呆改善建议系统 MVP 已启动：http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
